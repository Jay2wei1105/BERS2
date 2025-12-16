# -*- coding: utf-8 -*-
import openpyxl
import json

# 读取 Excel 文件
wb = openpyxl.load_workbook('建築能效計算表單.xlsm', data_only=True)

# 要提取的工作表名称
target_sheets = [
    "分區計算參數",
    "免評估分區計算參數", 
    "城鄉係數",
    "SORi"
]

result = {}

def extract_sheet_data(sheet_name, max_rows=1000):
    """提取工作表数据，返回清理后的数组"""
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ 工作表 '{sheet_name}' 不存在")
        return []
    
    ws = wb[sheet_name]
    data = []
    
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        # 跳过完全空白的行
        if not any(cell is not None and str(cell).strip() != '' for cell in row):
            continue
        
        # 转换为字符串数组
        row_data = []
        for cell in row:
            if cell is not None:
                # 处理数字类型
                if isinstance(cell, (int, float)):
                    row_data.append(cell)
                else:
                    row_data.append(str(cell).strip())
            else:
                row_data.append(None)
        
        data.append(row_data)
        
        if i >= max_rows:
            break
    
    return data

# 提取所有目标工作表
for sheet_name in target_sheets:
    print(f"📊 正在提取: {sheet_name}...")
    sheet_data = extract_sheet_data(sheet_name, max_rows=2000)
    result[sheet_name] = sheet_data
    print(f"   ✓ 已提取 {len(sheet_data)} 行数据")

# 保存为 JSON
with open('lookup_tables.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"\n✅ 完成！已导出到 lookup_tables.json")
print(f"📦 总共提取了 {len(result)} 个查找表")

# 显示每个表的预览
for sheet_name, data in result.items():
    print(f"\n{'='*60}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*60}")
    print(f"总行数: {len(data)}")
    if len(data) > 0:
        print(f"\n前 3 行预览:")
        for i, row in enumerate(data[:3], 1):
            # 只显示前10列
            preview = [str(cell)[:20] if cell is not None else '-' for cell in row[:10]]
            print(f"  行 {i}: {preview}")
